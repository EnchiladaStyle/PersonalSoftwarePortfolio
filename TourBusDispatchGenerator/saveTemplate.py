import sqlite3
from openpyxl import load_workbook

def create_connection(db_file):

    conn = None
    try:
        conn = sqlite3.connect(db_file)
        print("connected")
    except sqlite3.Error as e:
        print(e)
    return conn
    

def create_table(conn):

    create_table_sql = """CREATE TABLE IF NOT EXISTS templateModel 
    (id INTEGER PRIMARY KEY AUTOINCREMENT,
    dataSheet TEXT,
    toursAndLocations TEXT,
    distanceMatrix TEXT,
    name TEXT
    )"""

    try:
        with conn:
            cursor = conn.cursor()
            cursor.execute(create_table_sql)
            cursor.close()
    except sqlite3.Error as e:
        print(e)
    
def insert(conn, data):
    InsertSql = """INSERT INTO templateModel(dataSheet, toursAndLocations, distanceMatrix, name)
    VALUES(?, ?, ?, ?)"""

    with conn:
        cursor = conn.cursor()
        request = cursor.execute(InsertSql, data)
        conn.commit()
        cursor.close()

def deleteTemplateFromDatabase(id):
    conn = create_connection("test.db")
    with conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM templateModel WHERE id=?", (id,))
        conn.commit()
        cursor.close()

def getWorksheetFromExcel(wb, workSheetName):
    dataList = []
    countSinceData = 0
    for column in wb[workSheetName].iter_cols():

        dataList.append([])
        
        for cell in column:         

            if cell.value != None:
                countSinceData = 0
            elif cell.value == None:
                countSinceData += 1
            
            if (column[0].value == "pick up time" or column[0].value == "drop off time") and cell.value != "pick up time" and cell.value != "drop off time":
                if len(str(cell.value).split(":")) == 3:
                    dataList[-1].append(str(cell.value)[:-3])
                else:
                    dataList[-1].append(cell.value)
            else:
                dataList[-1].append(cell.value)
        if countSinceData > 3:
            dataList[-1] = dataList[-1][:1-countSinceData]        
        dataList[-1].append(None)
    dataString = ""
    for column in dataList:
        dataString += "@"
        for cell in column:
            dataString += f"{cell}"
            dataString += "~"
    dataString = dataString[1:-1]
    return dataString


def saveTemplate(excelFile, templateName):

    conn = create_connection("test.db")
    wb = load_workbook(excelFile)

    dataSheetString = getWorksheetFromExcel(wb, "Data Sheet")
    ToursAndLocationsString = getWorksheetFromExcel(wb, "Tours and Locations")
    DistanceMatrixString = getWorksheetFromExcel(wb, "Distance Matrix")
    wb.close()

    with conn:
        print(dataSheetString[:10])
        insert(conn, [dataSheetString, ToursAndLocationsString, DistanceMatrixString, templateName])

