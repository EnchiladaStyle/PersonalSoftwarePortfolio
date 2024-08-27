from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side


def styleTemplate(excelFile):

    wb = load_workbook(excelFile)

    DataSheet = wb["Data Sheet"]
    DataSheet.column_dimensions["A"].width = 20
    DataSheet.column_dimensions["B"].width = 38
    DataSheet.column_dimensions["C"].width = 12
    DataSheet.column_dimensions["D"].width = 20
    DataSheet.column_dimensions["E"].width = 21
    DataSheet.column_dimensions["F"].width = 12
    DataSheet.column_dimensions["G"].width = 13
    DataSheet.column_dimensions["H"].width = 14
    DataSheet.column_dimensions["I"].width = 10
    DataSheet.column_dimensions["J"].width = 17
    DataSheet.column_dimensions["K"].width = 14
    DataSheet.column_dimensions["L"].width = 13
    DataSheet.column_dimensions["M"].width = 10
    DataSheet.column_dimensions["N"].width = 15
    DataSheet.column_dimensions["O"].width = 15
    DataSheet.column_dimensions["P"].width = 16
    DataSheet.column_dimensions["Q"].width = 15
    DataSheet.column_dimensions["R"].width = 8

    DataSheet["A1"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    DataSheet["B1"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    DataSheet["C1"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    DataSheet["F1"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    DataSheet["G1"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    DataSheet["H1"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    DataSheet["J1"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    DataSheet["K1"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    DataSheet["L1"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    DataSheet["P1"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    DataSheet["Q1"].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    DataSheet["D1"].fill = PatternFill(start_color="D3D3D3", end_color="CCFFCC", fill_type="solid")
    DataSheet["E1"].fill = PatternFill(start_color="D3D3D3", end_color="CCFFCC", fill_type="solid")
    DataSheet["N1"].fill = PatternFill(start_color="D3D3D3", end_color="CCFFCC", fill_type="solid")
    DataSheet["O1"].fill = PatternFill(start_color="D3D3D3", end_color="CCFFCC", fill_type="solid")
    DataSheet["R1"].fill = PatternFill(start_color="D3D3D3", end_color="CCFFCC", fill_type="solid")

    counter = 2
    while counter < 200:
        DataSheet[f"F{counter}"].number_format = 'hh:mm'
        DataSheet[f"G{counter}"].number_format = 'hh:mm'
        counter += 1

    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    for col in DataSheet.iter_cols():
        for cell in col:
            cell.border = border
            break

    counter = 1
    distanceMatrix = wb["Distance Matrix"]
    for col in distanceMatrix.iter_cols():
        for cell in col:
            cell.border = border
            distanceMatrix.column_dimensions[cell.column_letter].width = 14
            cell.fill = PatternFill(start_color="D3D3D3", end_color="CCFFCC", fill_type="solid")
            distanceMatrix[f"A{counter}"].fill = PatternFill(start_color="D3D3D3", end_color="CCFFCC", fill_type="solid")
            distanceMatrix[f"A{counter}"].border = border
            counter += 1
            break

    ToursAndLocations = wb["Tours and Locations"]
    ToursAndLocations.column_dimensions["A"].width = 43
    ToursAndLocations.column_dimensions["B"].width = 26
    ToursAndLocations["A1"].fill = PatternFill(start_color="D3D3D3", end_color="CCFFCC", fill_type="solid")
    ToursAndLocations["B1"].fill = PatternFill(start_color="D3D3D3", end_color="CCFFCC", fill_type="solid")

    wb.save(excelFile)
    wb.close()