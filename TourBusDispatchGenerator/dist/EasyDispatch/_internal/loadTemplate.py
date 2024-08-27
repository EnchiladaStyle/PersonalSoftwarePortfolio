import sqlite3
from openpyxl import load_workbook
from styleTemplate import styleTemplate
from datetime import datetime
from saveTemplate import create_table

def selectTemplates():
    conn = create_connection("test.db")
    with conn:
        create_table(conn)
        cursor = conn.cursor()
        cursor.execute("SELECT name, id FROM templateModel")
        templates = cursor.fetchall()
        cursor.close

    return templates

def create_connection(db_file):

    conn = None
    try:
        conn = sqlite3.connect(db_file)
        print("connected")
    except sqlite3.Error as e:
        print(e)
    return conn

def getTemplateFromSQL(conn, id):

    with conn:
        cursor = conn.cursor()
        cursor.execute("SELECT dataSheet, toursAndLocations, distanceMatrix FROM templateModel WHERE id = ?", [id,])
        rows = cursor.fetchall()
        cursor.close()

    return rows[0]

def parseStringToLists(dataString):
    
    solution = []

    for i in range(len(dataString.split("@"))):
        solution.append(dataString.split("@")[i].split("~"))

    return solution

def isTime(string):
    formats = ['%H:%M', '%I:%M%p', '%I:%M %p']
    for fmt in formats:
        try:
            datetime.strptime(string, fmt)
            return True
        except ValueError:
            continue
    return False

def populateWorkSheet(sheet, newDataList):

    for col_idx, col_data in enumerate(newDataList, start=1):
        for row_idx, value in enumerate(col_data, start=1):
            if value == "None":
                sheet.cell(row=row_idx, column=col_idx, value=None)
            elif value.isdigit():
                sheet.cell(row=row_idx, column=col_idx, value=int(value)) 
            elif isTime(value):
                 sheet.cell(row=row_idx, column=col_idx, value=datetime.strptime(value, '%H:%M'))
            else:
                sheet.cell(row=row_idx, column=col_idx, value=value)


def loadTemplate(excelFile, id):

    wb = load_workbook(excelFile)
    conn = create_connection("test.db")

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    print("sheet must be empty")
                    return False
    
    for sheet in wb.worksheets:
        wb.remove(sheet)

    with conn:
        workSheetStrings = getTemplateFromSQL(conn, id)

    newDataSheetList = parseStringToLists(workSheetStrings[0])
    newToursAndLocationsList = parseStringToLists(workSheetStrings[1])
    newDistanceMatrixList = parseStringToLists(workSheetStrings[2])
    
    DataSheet = wb.create_sheet("Data Sheet")
    ToursAndLocations = wb.create_sheet("Tours and Locations")
    DistanceMatrix = wb.create_sheet("Distance Matrix")

    populateWorkSheet(DataSheet, newDataSheetList)
    populateWorkSheet(ToursAndLocations, newToursAndLocationsList)
    populateWorkSheet(DistanceMatrix, newDistanceMatrixList)

    wb.save(excelFile)
    wb.close()

    styleTemplate(excelFile)
    