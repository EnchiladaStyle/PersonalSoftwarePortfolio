
from openpyxl import load_workbook
from createDistanceMatrix import createDistanceMatrix

def createDataModel(excelFile):

    print("opening workbook in CreateDataModel")
    wb = load_workbook(excelFile, data_only=True)

    ws = wb["Data Sheet"]

    data = {}
    dockRepMatches = {}
    dockReps = []
    distanceMatrix = []
    locations = []
    times = []
    originalTimes = []
    vehicles = []
    vehicleCapacities = []
    passengerCounts = []
    drivers = []
    ships = []
    tourNames = []
    halfway = 0
    
    counter = 2
    for cell in ws["D"][1:]:
        if cell.value is not None:
            locations.append(cell.value)
            ships.append(ws[f"A{counter}"].value)
            tourNames.append(ws[f"B{counter}"].value)
            halfway += 1
        counter += 1

    counter = 2 
    for cell in ws["E"][1:]:
        if cell.value is not None:
            locations.append(cell.value)
            ships.append(ws[f"A{counter}"].value)
            tourNames.append(ws[f"B{counter}"].value)
        counter += 1
        
    locations.append(ws["R2"].value)
    for cell in ws["F"][1:]:
        if cell.value is not None:
            times.append((cell.value.hour*60 + cell.value.minute)//5)
            originalTimes.append(cell.value)
    for cell in ws["G"][1:]:
        if cell.value is not None:
            times.append((cell.value.hour*60 + cell.value.minute)//5)
            originalTimes.append(cell.value)
    originalTimes.append(0)
    counter = 2
    
    for cell in ws["N"][1:]:
        if cell.value is not None and ws[f"P{counter}"].value != "y":
            vehicles.append(cell.value)
            drivers.append(ws[f"Q{counter}"].value)
        counter += 1
    counter = 2
    for cell in ws["O"][1:]:
        if cell.value is not None and ws[f"P{counter}"].value != "y":
            vehicleCapacities.append(cell.value)
        counter += 1
    for cell in ws["H"][1:]:
        if cell.value is not None:
            passengerCounts.append(cell.value)
    for cell in ws["H"][1:]:
        if cell.value is not None:
            passengerCounts.append(cell.value * -1)

    counter = 2
    for cell in ws["L"][1:]:
        if cell.value is not None:
            dockRepMatches[ws[f"J{counter}"].value] = cell.value
        counter += 1
    for ship in ships:
        dockReps.append(dockRepMatches[ship])

    TravelTimes = createDistanceMatrix(wb['Distance Matrix'])
    
    row = []
    for location in locations:
        row = []
        for location2 in locations:
            row.append(TravelTimes[location][location2]//5)
        distanceMatrix.append(row)

    pickups_deliveries = []
    i = 0
    while i < halfway:
        pickups_deliveries.append((i, i+halfway))
        i += 1
        
    time_windows = []
    for time in times:
        time_windows.append((time - 1, time + 1))
    time_windows.append((0, 288))

    wb.close()
    
    if distanceMatrix == []:
        data["status"] = "No Distance Matrix Provided"
    elif locations == []:
        data["status"] = "No Locations Provided"
    elif pickups_deliveries == [] or time_windows == []:
        data["status"] = "Missing Data In Request"
    elif  vehicles == []:
        data["status"] = "No Vehicles Available"
    elif passengerCounts == []:
        data["status"] = "Passenger Counts Required"
    elif vehicleCapacities == []:
        data["status"] = "Vehicles Must Have Capacities"
    else:
        data["status"] = "Ready"

    data["time_matrix"] = distanceMatrix
    data["locations"] = locations
    data["pickups_deliveries"] = pickups_deliveries
    data["time_windows"] = time_windows
    data["vehicles"] = vehicles
    data["num_vehicles"] = len(vehicles)
    data["depot"] = len(data["locations"]) - 1
    data["original_times"] = originalTimes
    data["passenger_counts"] = passengerCounts
    data["vehicle_capacities"] = vehicleCapacities
    data["drivers"] = drivers
    data["shipNames"] = ships
    data["tourNames"] = tourNames
    data["dockReps"] = dockReps

    return data


