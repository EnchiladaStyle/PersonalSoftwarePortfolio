from openpyxl import load_workbook
from styleTemplate import styleTemplate


vehicleNumbers = [1, 2, 4, 6, 7, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 20, 21, 22, 23, 24, 25, 27, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 62, 63, 64, 65, 66]

vehicleCapacities = [20, 14, 14, 14, 14, 14, 14, 14, 14, 14, 11, 14, 14, 14, 3, 14, 14, 14, 14, 14, 14, 14, 14, 14, 11, 11, 14, 28, 2, 5, 14, 35, 35, 35, 35, 32, 24, 14, 14, 14, 14, 14, 14, 20, 20, 20, 20, 20, 20, 20, 20, 3, 50, 50, 50, 50, 50, 50, 50, 50]


def generateTemplate(excelFile):
    wb = load_workbook(excelFile)

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    print("sheet must be empty")
                    return False
    
    for sheet in wb.worksheets:
        wb.remove(sheet)
    
    DataSheet = wb.create_sheet("Data Sheet")
    ToursAndLocations = wb.create_sheet("Tours and Locations")
    DistanceMatrix = wb.create_sheet("Distance Matrix")

    #Data Sheet 
    DataSheet["A1"].value = "Ship Name"
    DataSheet["B1"].value = "Tour Name"
    DataSheet["C1"].value = "return? y/n"
    DataSheet["D1"].value = "pick up location"
    DataSheet["E1"].value = "drop off location"
    DataSheet["F1"].value = "pick up time"
    DataSheet["G1"].value = "drop off time"
    DataSheet["H1"].value = "passenger count"
    DataSheet["J1"].value = "ship names"
    DataSheet["K1"].value = "ship locations"
    DataSheet["L1"].value = "dock reps"
    DataSheet["N1"].value = "vehicle numbers"
    DataSheet["O1"].value = "vehicle capacities"
    DataSheet["P1"].value = "out of service? y/n"
    DataSheet["Q1"].value = "drivers"
    DataSheet["R1"].value = "base"
    DataSheet["R2"].value = "Cambria"



    counter = 2
    
    while counter < 200:
        DataSheet[f"D{counter}"].value = f"=IF(C{counter}=\"y\", IFERROR(INDEX(\'Tours and Locations\'!$B$1:$B$100, MATCH(B{counter}, \'Tours and Locations\'!$A$1:$A$100, 0)), \"\"), IFERROR(INDEX($K$1:$K$100, MATCH(A{counter}, $J$1:$J$100, 0)), \"\") )"
        DataSheet[f"E{counter}"].value = f"=IF(C{counter}<>\"y\", IFERROR(INDEX(\'Tours and Locations\'!$B$1:$B$100, MATCH(B{counter}, \'Tours and Locations\'!$A$1:$A$100, 0)), \"\"), IFERROR(INDEX($K$1:$K$100, MATCH(A{counter}, $J$1:$J$100, 0)), \"\"))"

        DataSheet[f"F{counter}"].value = f"=IF(C{counter}=\"y\", IFERROR(G{counter} - TIME(0, IFERROR(INDEX(\'Distance Matrix\'!$B$2:$AN$40, MATCH(D{counter}, \'Distance Matrix\'!$A$2:$A$40, 0), MATCH(E{counter}, \'Distance Matrix\'!$B$1:$AN$1, 0)), INDEX(\'Distance Matrix\'!$B$2:$AN$40, MATCH(E{counter}, \'Distance Matrix\'!$A$2:$A$40, 0), MATCH(D{counter}, \'Distance Matrix\'!$B$1:$AN$1, 0))), 0), \"\"), \"\")"
        DataSheet[f"G{counter}"].value = f"=IF(C{counter}<>\"y\", IFERROR(F{counter} + TIME(0, IFERROR(INDEX(\'Distance Matrix\'!$B$2:$AN$40, MATCH(D{counter}, \'Distance Matrix\'!$A$2:$A$40, 0), MATCH(E{counter}, \'Distance Matrix\'!$B$1:$AN$1, 0)), INDEX(\'Distance Matrix\'!$B$2:$AN$40, MATCH(E{counter}, \'Distance Matrix\'!$A$2:$A$40, 0), MATCH(D{counter}, \'Distance Matrix\'!$B$1:$AN$1, 0))), 0), \"\"), \"\")"
        counter += 1

    counter = 2
    while counter <= len(vehicleNumbers):
        DataSheet[f"N{counter}"].value = vehicleNumbers[counter-2]
        DataSheet[f"O{counter}"].value = vehicleCapacities[counter-2]
        DataSheet[f"P{counter}"].value = "y"
        counter += 1

    #Tours and Locations
    ToursAndLocationsData = [["Tour Names", "Drop off Locations"], ["Adventure Karts - Whipple Creek", "Whipple Creek"], ["Aerial Zip & Rappel - Wood Rd.", "Wood Road"],
    ["AK Duck Shuttle", "Front Street Extension"], ["AK Family Fun Exclusive", "Front Street Extension"], 
    ["AK First City Highlights & LJ Show", "Front Street Extension"], ["AK Lodge Adv & Seafeast - Clover Pass Resort", "Clover Pass"], ["Aleutian Ballad - Dock 3", "Dock 3"], 
    ["Baranof Fishing Excursions - Disco Center", "Disco"], ["Best of Ketchikan George Inlet Cannery", "Cannery"], 
    ["Cannery, Crab & Beer Flight - Cannery", "Cannery"], ["Cannery, Crab & Beer Flight - GIL", "George Inlet"], 
    ["Cannery, Small Bites & Saxman - Cannery", "Cannery"], ["Cannery, Small Bites & Saxman - Saxman", "Saxman"], 
    ["Canoes & Nature Trail - Harriet Hunt Lake", "Harriot Hunt"], ["Crab Cake Expedition - GIL", "George Inlet"], 
    ["Culinary, Cultural & LJ Show", "Saxman"], ["Eagle Island Kayaks - Potter Rd.", "Potter Road"], 
    ["George Inlet Lodge - Crab Feast", "George Inlet"], ["GIL Brunch & Saxman Showcase - GIL", "George Inlet"], 
    ["GIL Brunch & Saxman Showcase - Saxman", "Saxman"], ["Historic Cannery & LJ Show - Cannery", "Cannery"], 
    ["Historic Cannery & LJ Show - Disco Center", "Disco"], ["Historic Cannery & LJ Show - Front St. or Dock 1", "Dock 1"], 
    ["Hovercraft Shuttle", "Dock 3"], ["Jeeps & Canoes - Harriet Hunt Lake", "Harriot Hunt"], 
    ["Jeeps & Canoes - Warehouse", "Jeep Base"], ["Ketchikan Highlights by Trolley", "Saxman"], ["KIA - Gravina Side", "Airport"], 
    ["KIA - Ketchikan Side", "Airport"], ["Knudson Cove Sportfishing - Knudson Cove Marina", "Knudson"], 
    ["KTN Native Walking Tour - CFL", "Cape Fox Lodge"], ["KTN Native Walking Tour - Dock 4", "Dock 4"], 
    ["KTN Native Walking Tour - Salmon Ladder", "Salmon Ladder"], ["KTN Trolley & LJ Show", "Front Street Extension"], 
    ["Lighthouse, Eagles & Totems - Ward Cove", "Lighthouse"], ["LJ Show & Crab Feast - Front St. or Dock 1", "Dock 1"], 
    ["LJ Show &Crab Feast - GIL", "George Inlet"], ["Lumberjack Show - Disco Center", "Disco"], 
    ["Misty Fjords Flightseeing - Taquan Air", "Taquan"], ["Neets Bay Bear Cruise - Clover Pass Resort", "Clover Pass"], 
    ["Out to Sea Expedition - Thomas St.", "Thomas Street"], ["Rainforest Island Adv. - Knudson Cove Marina", "Knudson"], 
    ["RF Sanctuary Walk - Wood Rd.", "Wood Road"], ["RF Walk & Crab Feast - ARS to GIL", "George Inlet"], 
    ["RF Walk & Crab Feast - GIL", "George Inlet"], ["RF Walk & Crab Feast - GIL to ARS", "George Inlet"], 
    ["RF Walk & Crab Feast - Wood Rd.", "Wood Road"], ["RF Zip & Skybridge - Wood Rd.", "Wood Road"], 
    ["Ropes & Challenge Zips - Potter Rd.", "Potter Road"], ["Salmon Dip Making & Photostop", "Saxman"], 
    ["Salmon Dip Making & Photostop - Parking Lot", "Saxman"], ["Sawmill & Crab Feast - GIL", "George Inlet"], 
    ["Sawmill & Crab Feast - Parking Lot", "Saxman"], ["Sawmill & Crab Feast - Sawmill", "Saxman"], 
    ["Sawmill, Totems, & Discovery Center - Disco Center", "Disco"], ["Sawmill, Totems, & Discovery Center - Parking Lot", "Saxman"], 
    ["Sawmill, Totems, & Discovery Center - Sawmill", "Saxman"], ["Saxman Craft Showcase - Community Center", "Saxman"], 
    ["Saxman Craft Showcase - Parking Lot", "Saxman"], ["Saxman Native Village", "Saxman"], ["Saxman Native Village & LJ Show", "Saxman"], 
    ["Saxman Photostop & Crab Feast - GIL", "George Inlet"], ["Saxman Photostop & Crab Feast - Saxman", "Saxman"], 
    ["Tatoosh Island Kayaks - Potter Rd.", "Potter Road"], ["UTV Safari & GIL Crab Meal - GIL", "George Inlet"], 
    ["UTV Safari & GIL Crab Meal - White River", "White River"], ["White River UTV Base Camp", "White River"], 
    ["Wilderness Zodiac Quest - Cannery", "Cannery"], ["Wildlife Bear Country - Wood Rd.", "Wood Road"]]
    
    for nameAndLocation in ToursAndLocationsData:
        ToursAndLocations.append(nameAndLocation)

    #Distance Matrix
    DistanceMatrixData = [["Distance Matrix", "Ward Cove", "Dock 1", "Dock 2", "Dock  3", "Dock  4", "Front Street Extension", "George Inlet", "Cannery", "Wood Road", "Cape Fox Lodge", "Saxman", "Disco", "Taquan", "Jeep Base", "Harriot Hunt", "White River", "Whipple Creek", "Knudson", "Clover Pass", "Totem Bight", "Lighthouse", "Cambria", "Potter Road", "Airport", "Salmon Ladder", "Thomas Street"],
    ["Ward Cove",	0,	30,	30,	30,	30,	30,	45,	45,	40,	30,	30,	30,	15,	5,	30,	30,	20,	20,	20,	15,	5,	15,	30,	15,	30,	30],
    ["Dock 1", 30,	0,	5,	5,	10,	5,	30,	30,	20,	10,	15,	5,	15,	25,	45,	45,	30,	30,	30,	25,	30,	15,	30,	15,	15,	10],
    ["Dock 2", 30,	5,	0,	5,	10,	5,	30,	30,	20,	10,	15,	5,	15,	25,	45,	45,	30,	30,	30,	25,	25,	15,	30,	15,	15,	10],
    ["Dock 3", 30,	5,	5,	0,	10,	5,	30,	30,	20,	10,	15,	5,	15,	25,	45,	45,	30,	30,	30,	25,	25,	15,	30,	15,	15,	10],
    ["Dock 4", 30,	10,	10,	10,	0,	5,	30,	30,	20,	10,	15,	5,	15,	20,	45,	45,	30,	30,	30,	25,	25,	15,	30,	15,	15,	10],
    ["Front Street Extension",	30,	5,	5,	5,	5,	0,	30,	30,	20,	15,	15,	5,	15,	25,	45,	45,	30,	30,	30,	30,	30,	15,	30,	15,	15,	5],
    ["George Inlet",	45,	30,	30,	30,	30,	30,	0,	10,	20,	30,	15,	30,	45,	45,	75,	75,	60,	60,	60,	55,	60,	45,	60,	30,	30,	30],
    ["Cannery",	45,	30,	30,	30,	30,	30,	10,	0,	15,	30,	20,	30,	45,	60,	75,	75,	60,	75,	75,	70,	60,	45,	75,	45,	30,	30],
    ["Wood Road",	40,	20,	20,	20,	20,	20,	20,	15,	0,	30,	15,	20,	35,	40,	65,	65,	50,	50,	50,	50,	45,	35,	50,	40,	30,	30],
    ["Cape Fox Lodge",	30,	10,	10,	10,	10,	15,	30,	30,	30,	0,	15,	5,	15,	30,	45,	45,	30,	30,	30,	25,	30,	15,	30,	15,	5,	5,],
    ["Saxman",	30,	15,	15,	15,	15,	15,	15,	20,	15,	15,	0,	15,	15,	30,	60,	60,	45,	45,	45,	40,	55,	30,	45,	25,	10,	10],
    ["Disco",	30,	5,	5,	5,	5,	5,	30,	30,	20,	5,	15,	0,	15,	25,	45,	45,	30,	30,	30,	30,	30,	15,	30,	15,	5,	5],
    ["Taquan",	15,	15,	15,	15,	15,	15,	45,	45,	35,	15,	15,	15,	0,	15,	40,	40,	25,	30,	30,	25,	15,	5,	30,	5,	15,	15],
    ["Jeep Base",	5,	25,	25,	25,	20,	25,	45,	60,	40,	30,	30,	25,	15,	0,	30,	30,	25,	30,	30,	25,	5,	10,	30,	10,	20,	20],
    ["Harriot Hunt",	30,	45,	45,	45,	45,	45,	75,	75,	65,	45,	60,	45,	40,	30,	0,	30,	45,	50,	50,	40,	30,	45,	50,	45,	45,	50],
    ["White River",	30,	45,	45,	45,	45,	45,	75,	75,	65,	45,	60,	45,	40,	30,	30,	0,	45,	50,	50,	40,	30,	45,	50,	45,	45,	50],
    ["Whipple Creek",	20,	30,	30,	30,	30,	30,	60,	60,	50,	30,	45,	30,	25,	25,	45,	45,	0,	15,	15,	10,	20,	25,	15,	25,	30,	35],
    ["Knudson",	20,	30,	30,	30,	30,	30,	60,	75,	50,	30,	45,	30,	30,	30,	50,	50,	15,	0, 10,	15,	20,	25,	5,	25,	30,	35],
    ["Clover Pass",	20,	30,	30,	30,	30,	30,	60,	75,	50,	30,	45,	30,	30,	30,	50,	50,	15,	10,	0,	15,	20,	25,	10,	25,	30,	35],
    ["Totem Bight",	15,	25,	25,	25,	25,	30,	55,	70,	50,	25,	40,	30,	25,	25,	40,	40,	10,	15,	15,	0,	15,	20,	15,	20,	30,	35],
    ["Lighthouse",	5,	30,	25,	25,	25,	30,	60,	60,	45,	30,	55,	30,	15,	5,	30,	30,	20,	20,	20,	15,	0,	15,	20,	15,	30,	30],
    ["Cambria",	15,	15,	15,	15,	15,	15,	45,	45,	35,	15,	30,	15,	5,	10,	45,	45,	25,	25,	25,	20,	15,	0,	30,	5,	15,	15],
    ["Potter Road",	30,	30,	30,	30,	30,	30,	60,	75,	50,	30,	45,	30,	30,	30,	50,	50,	15,	5,	10,	15,	20,	30,	0,	40,	55,	45],
    ["Airport",	15,	15,	15,	15,	15,	15,	30,	45,	40,	15,	25,	15,	5,	10,	45,	45,	25,	25,	25,	20,	15,	5,	25,	0,	20,	10],
    ["Salmon Ladder",	30,	15,	15,	15,	15,	15,	30,	30,	30,	5,	10,	5,	15,	20,	45,	45,	30,	30,	30,	30,	30,	15,	30,	15,	0,	15],
    ["Thomas Street",	30,	10,	10,	10,	10,	5,	30,	30,	30,	5,	10,	5,	15,	20,	50,	50,	35,	35,	35,	35,	30,	15,	30,	15,	10,	0]
    ]

    for line in DistanceMatrixData:
        DistanceMatrix.append(line)
    
    


    
    
    wb.save(excelFile)
    wb.close()

    styleTemplate(excelFile)

    print("Thank you for providing an empty sheet.")


